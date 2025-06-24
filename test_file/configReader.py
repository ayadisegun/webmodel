import openpyxl

excelfile = "C:\\Users\\Segun\\PycharmProjects\\model\\test_file\\file.xlsx"
sheet_name = "details"
workbook = openpyxl.load_workbook(excelfile)
sheet = workbook[sheet_name]
totalrows =  sheet.max_row
totalcols = sheet.max_column

print("total rows are: ", str(totalrows), "and total column are: ", str(totalcols))

#reading
# print(sheet.cell(row=2, column=1).value)
#
# for rows in range(1, totalrows+1):
# 	for cols in range(1, totalcols+1):
# 		print(sheet.cell(row=rows, column=cols).value, end="     ")
# 	print()

#Writing
# sheet.cell(row=2, column=1).value="segun"
# sheet.cell(row=1, column=6).value="C-password"
# sheet.cell(row=2, column=6).value="123456"
# sheet.cell(row=3, column=6).value="7453778"
# workbook.save(excelfile)
# print(" ")
# print(" ")
# for rows in range(1, totalrows+1):
# 	for cols in range(1, totalcols+1):
# 		print(sheet.cell(row=rows, column=cols).value, end="     ")
# 	print()


