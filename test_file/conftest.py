import configparser
import inspect
import logging
import os
import time
import requests
from selenium import webdriver
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.chrome import options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC
import pytest
from selenium.webdriver.chrome.options import Options
from configparser import ConfigParser
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl


conftest_dir = os.path.dirname(os.path.abspath(__file__))
config_filepath = os.path.join(conftest_dir, 'config.ini')
config = configparser.ConfigParser()
if not os.path.exists(config_filepath):  # Good practice to check if file exists
    raise FileNotFoundError(f"Configuration file 'config.ini' not found at: {config_filepath}")
try:
    config.read(config_filepath)
    print(f"DEBUG: Config loaded from: {config_filepath}")
    print(f"DEBUG: Sections: {config.sections()}")
except Exception as e:
    raise RuntimeError(f"Error reading config file {config_filepath}: {e}")


def readconfig(section, key):
    try:
        return config.get(section, key)
    except configparser.NoSectionError:
        # Provide helpful error message if section is missing
        raise configparser.NoSectionError(f"No section '{section}' found in config file: {config_filepath}. Available sections: {config.sections()}")
    except configparser.NoOptionError:
        # Provide helpful error message if key is missing
        raise configparser.NoOptionError(f"No option '{key}' found in section '{section}' in config file: {config_filepath}")


def pytest_addoption(parser):
    parser.addoption(
        "--browser_name", action="store", default="chrome", help="Specify browser: chrome, firefox, or edge"
    )


@pytest.fixture(scope="class")
def setup(request):
    driver = None  #
    browser_name = request.config.getoption("browser_name")
    print(f"Running tests on browser: {browser_name}")
    preferences = {"download.default_directory": readconfig("setup", "download_directory")}
    # # preferences = {"download.default_directory": (readconfig("setup", "download_directory"))}
    # ops = webdriver.ChromeOptions()
    # ops.add_argument("--ignore-certificate-errors")
    # ops.add_argument("--start-maximized")
    # ops.add_argument("--disable-notification")
    # # ops.add_argument("--incognito")  # to open in incognito mode
    # # ops.add_argument("headless")
    # ops.add_experimental_option("prefs", preferences)
    # driver = webdriver.Chrome(options=ops)

    # --------------
    browser_name = request.config.getoption("browser_name")
    if browser_name == "chrome":
        ops = webdriver.ChromeOptions()
        ops.add_argument("--ignore-certificate-errors")
        ops.add_argument("--start-maximized")
        ops.add_argument("--disable-notifications")
        ops.add_experimental_option("prefs", preferences)
        # ops.add_argument("headless") # to run driver in headless mode
        driver = webdriver.Chrome(options=ops)
    elif browser_name == "edge":
        ops = webdriver.EdgeOptions()
        ops.add_argument("--ignore-certificate-errors")
        # ops.add_argument("headless") # to run driver in headless mode
        ops.add_argument("--start-maximized")
        ops.add_argument("--disable-notifications")
        ops.add_experimental_option("prefs", preferences)
        driver = webdriver.Edge(options=ops)
    elif browser_name == "firefox":
        ops = webdriver.FirefoxOptions()
        ops.add_argument(
            "--ignore-certificate-errors")
        ops.add_argument("--start-maximized")
        ops.add_argument("--disable-notifications")
        ops.set_preference("browser.download.dir", readconfig("setup", "download_directory"))
        ops.set_preference("browser.download.folderList", 2)  # 0=desktop, 1=downloads, 2=custom location
        ops.set_preference("browser.download.useDownloadDir", True)
        ops.set_preference("browser.helperApps.neverAsk.saveToDisk",
                           "application/octet-stream")  # Common types to auto-download
        driver = webdriver.Firefox(options=ops)
    else:
        raise ValueError(f"Unsupported browser: {browser_name}. Please choose 'chrome', 'firefox', or 'edge'.")
    driver.implicitly_wait(10)
    driver.get(readconfig("setup", "url"))
    request.cls.driver = driver
    yield driver
    print(f"Quitting {browser_name} browser.")
    driver.close()
    driver.quit()


@pytest.mark.usefixtures("setup")
class Utils:
    def __init__(self, driver):
        self.driver = driver

    def selectoption_bytext(self, locator, text):
        optns = Select(self.locator)
        optns.select_by_visible_text(text)


def send_mail(sender_address, sender_pass, receiver_address, subject, mail_content, attach_file_name,
              ):
    # sender_pass = 'Selenium@234'

    # setting up the MIME
    message = MIMEMultipart()
    message['From'] = sender_address
    message['To'] = receiver_address
    message['Subject'] = subject

    # The subject line
    # The body and the attachments for the mail
    message.attach(MIMEText(mail_content, 'plain'))
    attach_file = open(attach_file_name, 'rb')  # Open the file as binary mode
    payload = MIMEBase('application', 'octate-stream')
    payload.set_payload((attach_file).read())
    encoders.encode_base64(payload)  # encode the attachment
    # add payload header with filename
    payload.add_header('Content-Disposition', 'attachment', filename=attach_file_name)
    message.attach(payload)

    # Create SMTP session for sending the mail
    session = smtplib.SMTP_SSL("smtp.gmail.com", 465)
    session.login(sender_address, sender_pass)  # login with mail_id and password
    text = message.as_string()
    session.sendmail(sender_address, receiver_address, text)
    session.quit()
    print('Mail Sent')


def read_data():
    file_path = readconfig("data", "file")
    sheet_name = readconfig("data", "sheet_name")
    print(f"Attempting to open file: '{file_path}'")
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
    except FileNotFoundError:
        pytest.fail(f"Excel file not found: {file_path}")
    except KeyError:
        pytest.fail(f"Sheet '{sheet_name}' not found in the Excel file")

    max_row = sheet.max_row
    max_col = sheet.max_column

    # Get headers from first row
    headers = [sheet.cell(row=1, column=col).value for col in range(1, max_col + 1)]
    print("headers are: ", headers)

    data_rows = []
    for row in range(2, max_row + 1):
        row_data = [sheet.cell(row=row, column=col).value for col in range(1, max_col + 1)]
        data_rows.append(tuple(row_data))
    print(data_rows)
    return data_rows


def set_data(file_path, sheet_name, rowNum, colNum, data):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        sheet.cell(row=rowNum, column=colNum).value = data
        workbook.save(file_path)


@pytest.mark.hookwrapper
@pytest.hookimpl(hookwrapper=True, tryfirst=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    rep = outcome.get_result()
    setattr(item, "rep_" + rep.when, rep)
    return rep


@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    """
    Extends the PyTest Plugin to take and embed screenshot in html report, whenever test fails.
    :param item: The test item object
    :param call: The call object from pytest (used to get outcome)
    """
    pytest_html = item.config.pluginmanager.getplugin('html')
    outcome = yield
    report = outcome.get_result()
    extra = getattr(report, 'extra', [])

    # Check if the test has failed or was skipped due to xfail
    if report.when == 'call' or report.when == "setup":
        xfail = hasattr(report, 'wasxfail')
        if (report.skipped and xfail) or (report.failed and not xfail):
            driver = None
            try:
                driver = item.funcargs.get('setup')
            except Exception as e:
                print(f"Could not retrieve driver from fixture: {e}")
                driver = None

            if driver:
                # Define a directory for screenshots
                screenshot_dir = "hook-screenshots"
                os.makedirs(screenshot_dir, exist_ok=True) # Create directory if it doesn't exist

                safe_nodeid = report.nodeid.replace("::", "__").replace("/", "_").replace("[", "_").replace("]", "_")
                file_name = os.path.join(screenshot_dir, f"{safe_nodeid}.png")

                try:
                    driver.save_screenshot(file_name)
                    print(f"Screenshot saved: {file_name}")

                    if file_name:
                        html = f'<div><img src="{file_name}" alt="screenshot" style="width:304px;height:228px;" ' \
                               f'onclick="window.open(this.src)" align="right"/></div>'
                        extra.append(pytest_html.extras.html(html))
                except Exception as e:
                    print(f"Failed to take screenshot: {e}")
            else:
                print("Driver not available for screenshot.")
        report.extra = extra


@pytest.fixture(scope="module")
def get_logger():
    """
    Pytest fixture to provide a configured logger instance to tests.
    The logger name will be the name of the test file/module that requests it.
    """
    logs_dir = "logs"
    os.makedirs(logs_dir, exist_ok=True)

    calling_module = inspect.getmodule(inspect.stack()[1][3])
    logger_name = calling_module.__name__ if calling_module else "pytest_logger"
    logger = logging.getLogger(logger_name)

    if not logger.handlers:
        log_file_path = os.path.join(logs_dir, f"{logger_name}.log")
        file_handler = logging.FileHandler(log_file_path)
        formatter = logging.Formatter("%(asctime)s : %(levelname)s : %(name)s : %(message)s")
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        logger.setLevel(logging.INFO)  # info, debug, warning, error, critical
        # to print logs to console (optional)
        console_handler = logging.StreamHandler()
        console_handler.setFormatter(formatter)
        logger.addHandler(console_handler)

    yield logger


@pytest.hookimpl(tryfirst=True, hookwrapper=True)  # Correct decorator placement
def pytest_runtest_makereport(item, call):  # Added 'call' argument for the hookwrapper
    """
    Extends the PyTest Plugin to take and embed screenshot in html report, whenever test fails.
    :param item: The test item object
    :param call: The call object from pytest (used to get outcome)
    """
    pytest_html = item.config.pluginmanager.getplugin('html')
    outcome = yield # Capture the outcome of the test (pass/fail/skip)
    report = outcome.get_result()
    extra = getattr(report, 'extra', [])

    # Check if the test has failed or was skipped due to xfail
    if report.when == 'call' or report.when == "setup": # Check if the failure happened during setup or call phase
        xfail = hasattr(report, 'wasxfail')
        if (report.skipped and xfail) or (report.failed and not xfail):
            # Attempt to get the driver from the test item's fixtures
            # This is the crucial part to access the driver instance that was used by the test
            driver = None
            try:
                # Access the driver from the 'setup' fixture.
                # 'item.funcargs' contains the arguments passed to the test function,
                # including fixtures.
                driver = item.funcargs.get('setup')
            except Exception as e:
                print(f"Could not retrieve driver from fixture: {e}")
                driver = None

            if driver:
                # Define a directory for screenshots
                screenshot_dir = "hook-screenshots"
                os.makedirs(screenshot_dir, exist_ok=True)  # Create directory if it doesn't exist
                # Create a unique filename for the screenshot
                # Replace invalid characters for filenames
                safe_nodeid = report.nodeid.replace("::", "__").replace("/", "_").replace("[", "_").replace("]", "_")
                file_name = os.path.join(screenshot_dir, f"{safe_nodeid}.png")

                try:
                    driver.save_screenshot(file_name)  # Use driver.save_screenshot directly
                    print(f"Screenshot saved: {file_name}")

                    # Add the screenshot to the HTML report
                    if file_name:
                        # Construct the HTML for embedding the image in the report
                        html = f'<div><img src="{file_name}" alt="screenshot" style="width:304px;height:228px;" ' \
                               f'onclick="window.open(this.src)" align="right"/></div>'
                        extra.append(pytest_html.extras.html(html))
                except Exception as e:
                    print(f"Failed to take screenshot: {e}")
            else:
                print("Driver not available for screenshot.")
        report.extra = extra

